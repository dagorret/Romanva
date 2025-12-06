"""
Configuración del Admin de Django para Moodle Data
"""
from django.contrib import admin
from django.shortcuts import render, redirect
from django.urls import path
from django.utils.html import format_html
from django.contrib import messages
from django.core.management import call_command
from django.http import HttpResponse
from .models import (
    ImportLog, Course, Category, Enrol, UserEnrolment, 
    User, Group, GroupMember, UserLastAccess, RoleAssignment, Context
)
import io


# Configuración global del admin
admin.site.site_header = "Moodle Stats - Administración"
admin.site.site_title = "Moodle Stats Admin"
admin.site.index_title = "Panel de Control de Datos de Moodle"


@admin.register(ImportLog)
class ImportLogAdmin(admin.ModelAdmin):
    list_display = ['id', 'table_name', 'status_badge', 'records_imported', 'started_at', 'duration_display']
    list_filter = ['status', 'table_name', 'started_at']
    search_fields = ['table_name', 'error_message']
    readonly_fields = ['table_name', 'started_at', 'completed_at', 'records_imported', 'status', 'error_message']
    
    def status_badge(self, obj):
        colors = {
            'running': 'blue',
            'completed': 'green',
            'failed': 'red',
        }
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 10px; border-radius: 3px;">{}</span>',
            colors.get(obj.status, 'gray'),
            obj.get_status_display()
        )
    status_badge.short_description = 'Estado'
    
    def duration_display(self, obj):
        if obj.completed_at and obj.started_at:
            duration = obj.completed_at - obj.started_at
            total_seconds = int(duration.total_seconds())
            minutes, seconds = divmod(total_seconds, 60)
            return f"{minutes}m {seconds}s"
        return "-"
    duration_display.short_description = 'Duración'
    
    def has_add_permission(self, request):
        return False
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('export-all/', self.admin_site.admin_view(self.export_all_view), name='moodledata_export_all'),
            path('import-all/', self.admin_site.admin_view(self.import_all_view), name='moodledata_import_all'),
        ]
        return custom_urls + urls
    
    def export_all_view(self, request):
        """Exporta todas las tablas desde Moodle"""
        if request.method == 'POST':
            try:
                # Capturar output del comando
                out = io.StringIO()
                call_command('export_moodle', '--tables', 'all', stdout=out)
                
                messages.success(request, f'✓ Exportación completada exitosamente')
                messages.info(request, out.getvalue())
                
            except Exception as e:
                messages.error(request, f'Error en la exportación: {str(e)}')
            
            return redirect('admin:moodledata_importlog_changelist')
        
        # GET - Mostrar formulario de confirmación
        context = {
            'title': 'Exportar todas las tablas desde Moodle',
            'opts': self.model._meta,
            'has_view_permission': True,
            'tables': [
                'courses', 'categories', 'enrol', 'user_enrolments', 'users',
                'groups', 'groups_members', 'user_lastaccess', 'role_assignments', 'context'
            ]
        }
        return render(request, 'admin/export_all_confirm.html', context)
    
    def import_all_view(self, request):
        """Importa todas las tablas a Django desde archivos NDJSON"""
        if request.method == 'POST':
            clear = request.POST.get('clear') == 'on'
            
            try:
                # Capturar output del comando
                out = io.StringIO()
                if clear:
                    call_command('import_moodle', '--tables', 'all', '--clear', stdout=out)
                else:
                    call_command('import_moodle', '--tables', 'all', stdout=out)
                
                messages.success(request, f'✓ Importación completada exitosamente')
                messages.info(request, out.getvalue())
                
            except Exception as e:
                messages.error(request, f'Error en la importación: {str(e)}')
            
            return redirect('admin:moodledata_importlog_changelist')
        
        # GET - Mostrar formulario de confirmación
        context = {
            'title': 'Importar todas las tablas a Django',
            'opts': self.model._meta,
            'has_view_permission': True,
            'tables': [
                'courses', 'categories', 'enrol', 'user_enrolments', 'users',
                'groups', 'groups_members', 'user_lastaccess', 'role_assignments', 'context'
            ]
        }
        return render(request, 'admin/import_all_confirm.html', context)


# Action para exportar a Excel
def export_to_excel(modeladmin, request, queryset):
    """Exporta los registros seleccionados a Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from datetime import datetime
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = modeladmin.model._meta.verbose_name_plural[:31]
    
    # Obtener campos del modelo
    fields = [f for f in modeladmin.model._meta.fields if f.name not in ['id']]
    
    # Encabezados
    headers = [f.verbose_name for f in fields]
    ws.append(headers)
    
    # Estilo de encabezados
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Datos
    for obj in queryset:
        row = []
        for field in fields:
            value = getattr(obj, field.name)
            row.append(value)
        ws.append(row)
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Generar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{modeladmin.model._meta.verbose_name_plural}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response

export_to_excel.short_description = "📊 Exportar seleccionados a Excel"


# Admin para cada modelo de datos

@admin.register(Course)
class CourseAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'shortname', 'fullname', 'category', 'visible', 'imported_at']
    list_filter = ['visible', 'category', 'imported_at']
    search_fields = ['shortname', 'fullname']
    readonly_fields = ['moodle_id', 'imported_at']
    list_per_page = 50
    actions = [export_to_excel]


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'name', 'parent', 'path', 'visible', 'imported_at']
    list_filter = ['visible', 'parent', 'imported_at']
    search_fields = ['name', 'path']
    readonly_fields = ['moodle_id', 'imported_at']
    actions = [export_to_excel]


@admin.register(Enrol)
class EnrolAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'courseid', 'enrol', 'status', 'imported_at']
    list_filter = ['enrol', 'status', 'imported_at']
    search_fields = ['courseid']
    readonly_fields = ['moodle_id', 'imported_at']
    actions = [export_to_excel]


@admin.register(UserEnrolment)
class UserEnrolmentAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'userid', 'enrolid', 'status', 'imported_at']
    list_filter = ['status', 'imported_at']
    search_fields = ['userid', 'enrolid']
    readonly_fields = ['moodle_id', 'imported_at']
    list_per_page = 100
    actions = [export_to_excel]


@admin.register(User)
class UserAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'username', 'full_name', 'email', 'suspended', 'imported_at']
    list_filter = ['suspended', 'deleted', 'country', 'imported_at']
    search_fields = ['username', 'firstname', 'lastname', 'email']
    readonly_fields = ['moodle_id', 'imported_at']
    list_per_page = 50
    actions = [export_to_excel]
    
    def full_name(self, obj):
        return f"{obj.lastname}, {obj.firstname}"
    full_name.short_description = 'Nombre Completo'


@admin.register(Group)
class GroupAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'name', 'courseid', 'idnumber', 'imported_at']
    list_filter = ['courseid', 'imported_at']
    search_fields = ['name', 'idnumber']
    readonly_fields = ['moodle_id', 'imported_at']
    actions = [export_to_excel]


@admin.register(GroupMember)
class GroupMemberAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'groupid', 'userid', 'imported_at']
    list_filter = ['imported_at']
    search_fields = ['groupid', 'userid']
    readonly_fields = ['moodle_id', 'imported_at']
    list_per_page = 100
    actions = [export_to_excel]


@admin.register(UserLastAccess)
class UserLastAccessAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'userid', 'courseid', 'timeaccess', 'imported_at']
    list_filter = ['imported_at']
    search_fields = ['userid', 'courseid']
    readonly_fields = ['moodle_id', 'imported_at']
    list_per_page = 100
    actions = [export_to_excel]


@admin.register(RoleAssignment)
class RoleAssignmentAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'userid', 'roleid', 'contextid', 'timemodified', 'imported_at']
    list_filter = ['roleid', 'imported_at']
    search_fields = ['userid', 'contextid']
    readonly_fields = ['moodle_id', 'imported_at']
    list_per_page = 100
    actions = [export_to_excel]


@admin.register(Context)
class ContextAdmin(admin.ModelAdmin):
    list_display = ['moodle_id', 'contextlevel', 'instanceid', 'depth', 'imported_at']
    list_filter = ['contextlevel', 'depth', 'imported_at']
    search_fields = ['instanceid', 'path']
    readonly_fields = ['moodle_id', 'imported_at']
    list_per_page = 100
    actions = [export_to_excel]
